"""
Vistas de la API (ViewSets) para el sistema de Inventario TI.
Incluye endpoints CRUD, generacion de reportes (PDF/Excel),
historial de auditoria y endpoints estadisticos para Dashboards.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from django.http import HttpResponse

from rest_framework import viewsets, filters, serializers as drf_serializers
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

# Modelos
from .models import (
    Marca, Modelo, Procesador, Ram, Almacenamiento,
    SistemaOperativo, Software,
    Area, CentroCosto, Empleado,
    Computador, Celular, Chip, Monitor, ComputadorSoftware,
    PerfilUsuario, RolUsuario,
    SolicitudCorreo, EstadoSolicitudCorreo,
)
# Serializadores
from .serializers import (
    MarcaSerializer, ModeloSerializer, ProcesadorSerializer,
    RamSerializer, AlmacenamientoSerializer,
    SistemaOperativoSerializer, SoftwareSerializer,
    AreaSerializer, CentroCostoSerializer, EmpleadoSerializer,
    ComputadorReadSerializer, ComputadorWriteSerializer,
    CelularReadSerializer, CelularWriteSerializer,
    ChipReadSerializer, ChipWriteSerializer,
    MonitorReadSerializer, MonitorWriteSerializer,
    ComputadorSoftwareReadSerializer, ComputadorSoftwareWriteSerializer,
    UsuarioSerializer,
    SolicitudCorreoReadSerializer, SolicitudCorreoWriteSerializer,
)

User = get_user_model()


# ═══════════════════════════════════════════════
#  MIXIN: Filtrado por Centro de Costo según rol
# ═══════════════════════════════════════════════

class RolCCFilterMixin:
    """
    Mixin que restringe el queryset al Centro de Costo del usuario
    cuando su rol es ENCARGADO_OBRA.
    """
    # Subclasses override this to tell the mixin how to filter by CC id
    cc_filter_field = "empleado_asignado__centro_costo_id"

    def _get_cc_id(self):
        """
        Retorna el centro_costo_id del usuario si su rol es ENCARGADO_OBRA.
        Retorna None para administradores y viewers (sin restricción de CC).
        Maneja con seguridad el caso donde el perfil no existe.
        """
        user = self.request.user
        if not user.is_authenticated:
            return None
        try:
            perfil = user.perfil
            if perfil.rol == RolUsuario.ENCARGADO_OBRA and perfil.centro_costo_id:
                return perfil.centro_costo_id
        except PerfilUsuario.DoesNotExist:
            pass
        return None

    def get_queryset(self):
        """
        Filtra el queryset base para que ENCARGADO_OBRA solo vea equipos de su CC.
        Utiliza cc_filter_field para adaptarse a los distintos modelos
        (ej: 'empleado_asignado__centro_costo_id' para computadores, 'centro_costo_id' para monitores).
        """
        qs = super().get_queryset()
        cc_id = self._get_cc_id()
        if cc_id:
            qs = qs.filter(**{self.cc_filter_field: cc_id})
        return qs


# ═══════════════════════════════════════════════
#  CATALOGOS (Sin paginacion para listas desplegables rapidas)
# ═══════════════════════════════════════════════

class MarcaViewSet(viewsets.ModelViewSet):
    """
    Manejo del catálogo de Marcas.
    Acepta ?tipo_equipo=COMPUTADOR|CELULAR|MONITOR para devolver solo las marcas
    que tienen al menos un modelo de ese tipo.
    """
    queryset = Marca.objects.all()
    serializer_class = MarcaSerializer
    search_fields = ["nombre"]
    pagination_class = None

    def get_queryset(self):
        """Filtra marcas por tipo de equipo si se recibe ?tipo_equipo=COMPUTADOR|CELULAR|MONITOR."""
        qs = super().get_queryset()
        tipo = self.request.query_params.get("tipo_equipo")
        if tipo:
            qs = qs.filter(modelos__tipo_equipo=tipo).distinct()
        return qs


class ModeloViewSet(viewsets.ModelViewSet):
    """
    Manejo de Modelos, filtrables por Marca y por tipo_equipo.
    Acepta ?tipo_equipo=COMPUTADOR|CELULAR|MONITOR para filtrar por categoría.
    """
    queryset = Modelo.objects.select_related("marca").all()
    serializer_class = ModeloSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["marca", "tipo_equipo"]
    search_fields = ["nombre", "marca__nombre"]
    pagination_class = None


class ProcesadorViewSet(viewsets.ModelViewSet):
    """Catalogo de Procesadores."""
    queryset = Procesador.objects.all()
    serializer_class = ProcesadorSerializer
    search_fields = ["nombre"]
    pagination_class = None


class RamViewSet(viewsets.ModelViewSet):
    """Catalogo de capacidades RAM."""
    queryset = Ram.objects.all()
    serializer_class = RamSerializer
    search_fields = ["capacidad"]
    pagination_class = None


class AlmacenamientoViewSet(viewsets.ModelViewSet):
    """Catalogo de capacidades de almacenamiento."""
    queryset = Almacenamiento.objects.all()
    serializer_class = AlmacenamientoSerializer
    search_fields = ["capacidad", "nombre_modelo"]
    pagination_class = None


class SistemaOperativoViewSet(viewsets.ModelViewSet):
    """Catalogo de sistemas operativos."""
    queryset = SistemaOperativo.objects.all()
    serializer_class = SistemaOperativoSerializer
    search_fields = ["nombre"]
    pagination_class = None


class SoftwareViewSet(viewsets.ModelViewSet):
    """Catalogo de software."""
    queryset = Software.objects.all()
    serializer_class = SoftwareSerializer
    search_fields = ["nombre", "fabricante"]
    pagination_class = None


# ═══════════════════════════════════════════════
#  ORGANIZACION
# ═══════════════════════════════════════════════

class AreaViewSet(viewsets.ModelViewSet):
    """Gestion de las areas de la empresa."""
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    search_fields = ["nombre"]
    pagination_class = None


class EmpleadoViewSet(RolCCFilterMixin, viewsets.ModelViewSet):
    """Gestion de empleados para asignarles equipos."""
    queryset = Empleado.objects.select_related("area", "centro_costo").all()
    serializer_class = EmpleadoSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["area", "activo", "centro_costo"]
    search_fields = ["first_name", "last_name", "username", "numero_documento", "cargo"]
    pagination_class = None
    cc_filter_field = "centro_costo_id"


class CentroCostoViewSet(viewsets.ModelViewSet):
    """Gestion de centros de costo por area."""
    queryset = CentroCosto.objects.select_related("area").all()
    serializer_class = CentroCostoSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["area", "tipo", "activo"]
    search_fields = ["codigo", "nombre"]
    pagination_class = None

    @action(detail=True, methods=["get"])
    def resumen(self, request, pk=None):
        """
        Devuelve un resumen completo del CC: empleados, equipos por tipo y totales.
        GET /api/v1/centros-costo/{id}/resumen/
        """
        cc = self.get_object()
        empleados = Empleado.objects.filter(centro_costo=cc, activo=True).select_related("area")
        emp_ids = list(empleados.values_list("id", flat=True))

        computadores = Computador.objects.filter(empleado_asignado__in=emp_ids)
        celulares    = Celular.objects.filter(empleado_asignado__in=emp_ids)
        chips        = Chip.objects.filter(empleado_asignado__in=emp_ids)
        monitores    = Monitor.objects.filter(centro_costo=cc)

        def _eq_summary(qs, fields):
            return list(qs.values(*fields))

        # Equipos pendientes de devolución
        pendientes = (
            computadores.filter(estado="PENDIENTE_DEVOLUCION").count()
            + celulares.filter(estado="PENDIENTE_DEVOLUCION").count()
            + monitores.filter(estado="PENDIENTE_DEVOLUCION").count()
        )

        data = {
            "centro_costo": {
                "id": cc.id,
                "codigo": cc.codigo,
                "nombre": cc.nombre,
                "tipo": cc.tipo,
                "tipo_display": cc.get_tipo_display(),
                "area_nombre": cc.area.nombre if cc.area else "",
                "activo": cc.activo,
                "descripcion": cc.descripcion,
            },
            "totales": {
                "empleados":     empleados.count(),
                "computadores":  computadores.count(),
                "celulares":     celulares.count(),
                "chips":         chips.count(),
                "monitores":     monitores.count(),
                "pendientes_devolucion": pendientes,
            },
            "computadores_por_estado": dict(
                computadores.values_list("estado").annotate(c=Count("id")).values_list("estado", "c")
            ),
            "celulares_por_estado": dict(
                celulares.values_list("estado").annotate(c=Count("id")).values_list("estado", "c")
            ),
            "empleados_lista": [
                {
                    "id": e.id,
                    "nombre": e.get_full_name(),
                    "cargo": e.cargo,
                    "computadores": e.computadores.count(),
                    "celulares": e.celulares.count(),
                    "chips": e.chips.count(),
                }
                for e in empleados
            ],
            "equipos_pendientes": [
                *[
                    {
                        "tipo": "Computador",
                        "descripcion": str(eq),
                        "numero_inventario": eq.numero_inventario,
                        "asignado_a": eq.empleado_asignado.get_full_name() if eq.empleado_asignado else "",
                        "estado": eq.get_estado_display(),
                    }
                    for eq in computadores.filter(estado="PENDIENTE_DEVOLUCION")
                ],
                *[
                    {
                        "tipo": "Celular",
                        "descripcion": str(eq),
                        "numero_inventario": eq.numero_linea or eq.imei or str(eq.id),
                        "asignado_a": eq.empleado_asignado.get_full_name() if eq.empleado_asignado else "",
                        "estado": eq.get_estado_display(),
                    }
                    for eq in celulares.filter(estado="PENDIENTE_DEVOLUCION")
                ],
            ],
        }
        return Response(data)

    @action(detail=True, methods=["get"])
    def movimientos(self, request, pk=None):
        """
        Devuelve el historial de cambios de todos los equipos asignados al CC.
        GET /api/v1/centros-costo/{id}/movimientos/
        """
        cc = self.get_object()
        emp_ids = list(Empleado.objects.filter(centro_costo=cc).values_list("id", flat=True))

        results = []
        results += _serialize_history(
            Computador.history
            .filter(
                Q(empleado_asignado__in=emp_ids) | Q(empleado_asignado__isnull=True,
                    id__in=Computador.objects.filter(empleado_asignado__in=emp_ids).values("id"))
            )
            .select_related("history_user")
            .order_by("-history_date"),
            "Computador",
            limit=100,
        )
        results += _serialize_history(
            Celular.history
            .filter(empleado_asignado__in=emp_ids)
            .select_related("history_user")
            .order_by("-history_date"),
            "Celular",
            limit=100,
        )
        results.sort(key=lambda x: x["fecha"], reverse=True)
        return Response(results[:150])


# ═══════════════════════════════════════════════
#  USUARIOS (gestión de cuentas + roles)
# ═══════════════════════════════════════════════

class UsuarioViewSet(viewsets.ModelViewSet):
    """
    Gestión de usuarios del sistema.
    Permite crear, listar, editar y asignar roles + CC a cada usuario.
    """
    queryset = User.objects.prefetch_related("perfil", "perfil__centro_costo").all()
    serializer_class = UsuarioSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["username", "first_name", "last_name", "email"]
    pagination_class = None


class ChipViewSet(RolCCFilterMixin, viewsets.ModelViewSet):
    """Gestion de chips / SIMs (entidad independiente)."""
    queryset = Chip.objects.select_related("empleado_asignado", "celular").all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["estado", "operador", "empleado_asignado"]
    search_fields = ["numero_linea", "iccid", "plan"]
    pagination_class = None

    def get_serializer_class(self):
        """Usa el serializer de escritura (solo IDs) para mutaciones y el de lectura (objetos anidados) para GET."""
        if self.action in ("create", "update", "partial_update"):
            return ChipWriteSerializer
        return ChipReadSerializer


# ═══════════════════════════════════════════════
#  HELPERS — Exportacion de Reportes
# ═══════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

PDF_TABLE_STYLE = TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("FONTSIZE", (0, 0), (-1, 0), 9),
    ("FONTSIZE", (0, 1), (-1, -1), 7),
    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
    ("TOPPADDING", (0, 0), (-1, 0), 10),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
])

def _excel_response(ws, wb, filename):
    """
    Aplica estilos corporativos (fondo azul, fuente blanca) a la fila de cabecera del worksheet,
    ajusta el ancho de todas las columnas a 20 y retorna un HttpResponse para descarga Excel.
    """
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp

def _pdf_response(title, headers, rows, filename):
    """
    Genera un PDF en formato apaisado (landscape letter) con título y tabla estilizada.
    - title: texto del encabezado del documento
    - headers: lista de cadenas para la fila de cabecera de la tabla
    - rows: lista de listas con los datos de cada fila
    - filename: nombre sugerido para la descarga
    Retorna HttpResponse con Content-Type application/pdf.
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Spacer(1, 12),
    ]
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(PDF_TABLE_STYLE)
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def _asignado_nombre(obj):
    """Retorna el nombre completo del empleado asignado al equipo, o 'Sin asignar' si no tiene."""
    if obj.empleado_asignado:
        return obj.empleado_asignado.get_full_name()
    return "Sin asignar"

def _asignado_area(obj):
    """Retorna el nombre del área del empleado asignado al equipo, o cadena vacía si no aplica."""
    if obj.empleado_asignado and obj.empleado_asignado.area:
        return obj.empleado_asignado.area.nombre
    return ""


# ═══════════════════════════════════════════════
#  HELPERS — Historial
# ═══════════════════════════════════════════════

def _serialize_history(queryset, tipo_equipo, limit=100):
    """
    Convierte registros de django-simple-history en una lista de dicts JSON-friendly.
    Por cada registro calcula el diff contra la versión anterior para listar los campos modificados.
    - queryset: HistoricalQueryset del modelo (ej: Computador.history.all())
    - tipo_equipo: string identificador para el frontend (ej: 'Computador', 'Celular')
    - limit: máximo de registros a procesar (por defecto 100)
    Retorna lista de dicts con: id, tipo_equipo, equipo, equipo_id, fecha, tipo_cambio, usuario, cambios.
    """
    results = []
    for record in queryset[:limit]:
        cambios = []
        if record.prev_record:
            try:
                delta = record.diff_against(record.prev_record)
                cambios = [
                    {"campo": c.field, "anterior": str(c.old), "nuevo": str(c.new)}
                    for c in delta.changes
                ]
            except Exception:
                pass
        results.append({
            "id": record.history_id,
            "tipo_equipo": tipo_equipo,
            "equipo": str(record),
            "equipo_id": record.id,
            "fecha": record.history_date,
            "tipo_cambio": record.get_history_type_display(),
            "usuario": record.history_user.username if record.history_user else "Sistema",
            "cambios": cambios,
        })
    return results


# ═══════════════════════════════════════════════
#  EQUIPOS — COMPUTADOR
# ═══════════════════════════════════════════════

class ComputadorViewSet(RolCCFilterMixin, viewsets.ModelViewSet):
    """Gestion de Computadores (laptops, desktops, all-in-one, workstations)."""
    queryset = Computador.objects.select_related(
        "marca", "modelo", "modelo__marca", "procesador",
        "ram", "almacenamiento", "sistema_operativo",
        "empleado_asignado", "empleado_asignado__area",
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        "estado", "tipo_equipo", "marca", "modelo",
        "procesador", "empleado_asignado", "en_inventario",
    ]
    search_fields = [
        "numero_inventario", "numero_serie",
        "marca__nombre", "modelo__nombre",
    ]
    ordering_fields = ["fecha_ingreso", "estado", "marca__nombre"]

    def get_queryset(self):
        """
        Extiende el filtro del mixin de CC y agrega soporte para:
        ?con_empleado=true → solo equipos con empleado asignado (usado por el selector de Monitor).
        """
        qs = super().get_queryset()
        # ?con_empleado=true → solo equipos con empleado asignado (para selector en Monitor)
        if self.request.query_params.get("con_empleado") == "true":
            qs = qs.filter(empleado_asignado__isnull=False)
        return qs

    def get_serializer_class(self):
        """Usa ComputadorReadSerializer (anidado) para GET y ComputadorWriteSerializer (IDs planos) para escritura."""
        if self.action in ("list", "retrieve"):
            return ComputadorReadSerializer
        return ComputadorWriteSerializer

    @action(detail=True, methods=["get"])
    def historial(self, request, pk=None):
        """Devuelve el historial de cambios de un Computador específico con los diffs entre versiones."""
        computador = self.get_object()
        data = _serialize_history(computador.history.all(), "Computador")
        return Response(data)

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Genera y descarga un archivo Excel con todos los computadores del queryset filtrado."""
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Computadores"
        ws.append([
            "N Inventario", "N Serie", "Tipo", "Marca", "Modelo",
            "Procesador", "RAM", "Almacenamiento", "S.O.",
            "Estado", "Asignado a", "Area",
            "Fecha Ingreso", "Notas",
        ])
        for eq in qs:
            ws.append([
                eq.numero_inventario, eq.numero_serie,
                eq.get_tipo_equipo_display(),
                str(eq.marca), eq.modelo.nombre,
                str(eq.procesador) if eq.procesador else "",
                str(eq.ram) if eq.ram else "",
                str(eq.almacenamiento) if eq.almacenamiento else "",
                str(eq.sistema_operativo) if eq.sistema_operativo else "",
                eq.get_estado_display(), _asignado_nombre(eq),
                _asignado_area(eq), str(eq.fecha_ingreso), eq.notas,
            ])
        return _excel_response(ws, wb, "computadores_inventario.xlsx")

    @action(detail=False, methods=["get"], url_path="exportar-pdf")
    def exportar_pdf(self, request):
        """Genera y descarga un PDF apaisado con los computadores del queryset filtrado."""
        qs = self.filter_queryset(self.get_queryset())
        headers = [
            "N Inv.", "N Serie", "Tipo", "Marca", "Modelo",
            "Procesador", "RAM", "Disco", "Estado", "Asignado a",
        ]
        rows = [
            [
                eq.numero_inventario, eq.numero_serie,
                eq.get_tipo_equipo_display(),
                str(eq.marca), eq.modelo.nombre,
                str(eq.procesador) if eq.procesador else "",
                str(eq.ram) if eq.ram else "",
                str(eq.almacenamiento) if eq.almacenamiento else "",
                eq.get_estado_display(), _asignado_nombre(eq),
            ]
            for eq in qs
        ]
        return _pdf_response(
            "Reporte de Inventario — Computadores",
            headers, rows, "computadores_inventario.pdf",
        )


# ═══════════════════════════════════════════════
#  EQUIPOS — CELULAR
# ═══════════════════════════════════════════════

class CelularViewSet(RolCCFilterMixin, viewsets.ModelViewSet):
    """Gestion de Celulares y Tablets con exportacion e historial."""
    queryset = Celular.objects.select_related(
        "marca", "modelo", "modelo__marca",
        "ram", "almacenamiento", "empleado_asignado", "empleado_asignado__area",
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["estado", "tipo_equipo", "marca", "modelo", "empleado_asignado"]
    search_fields = [
        "numero_linea", "imei", "numero_serie",
        "marca__nombre", "modelo__nombre",
    ]
    ordering_fields = ["fecha_ingreso", "estado", "marca__nombre"]

    def get_queryset(self):
        """
        Extiende el filtro del mixin de CC y agrega soporte para:
        ?con_empleado=true → solo celulares con empleado asignado (usado por el selector de Chip).
        """
        qs = super().get_queryset()
        # ?con_empleado=true → solo celulares con empleado asignado (para selector en Chip)
        if self.request.query_params.get("con_empleado") == "true":
            qs = qs.filter(empleado_asignado__isnull=False)
        return qs

    def get_serializer_class(self):
        """Usa CelularReadSerializer (anidado) para GET y CelularWriteSerializer (IDs planos) para escritura."""
        if self.action in ("list", "retrieve"):
            return CelularReadSerializer
        return CelularWriteSerializer

    @action(detail=True, methods=["get"])
    def historial(self, request, pk=None):
        """Devuelve el historial de cambios de un Celular específico con los diffs entre versiones."""
        celular = self.get_object()
        data = _serialize_history(celular.history.all(), "Celular")
        return Response(data)

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Genera y descarga un archivo Excel con todos los celulares del queryset filtrado."""
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Celulares"
        ws.append([
            "N Linea", "IMEI", "N Serie", "Tipo", "Marca", "Modelo",
            "RAM", "Almacenamiento", "Estado", "Asignado a", "Area",
            "Fecha Ingreso", "Notas",
        ])
        for cl in qs:
            ws.append([
                cl.numero_linea or "", cl.imei or "", cl.numero_serie or "",
                cl.get_tipo_equipo_display(),
                str(cl.marca), cl.modelo.nombre,
                str(cl.ram) if cl.ram else "",
                str(cl.almacenamiento) if cl.almacenamiento else "",
                cl.get_estado_display(),
                _asignado_nombre(cl), _asignado_area(cl),
                str(cl.fecha_ingreso), cl.notas,
            ])
        return _excel_response(ws, wb, "celulares_inventario.xlsx")

    @action(detail=False, methods=["get"], url_path="exportar-pdf")
    def exportar_pdf(self, request):
        """Genera y descarga un PDF apaisado con los celulares del queryset filtrado."""
        qs = self.filter_queryset(self.get_queryset())
        headers = [
            "N Linea", "IMEI", "N Serie", "Marca", "Modelo",
            "RAM", "Disco", "Estado", "Asignado a",
        ]
        rows = [
            [
                cl.numero_linea or "", cl.imei or "", cl.numero_serie or "",
                str(cl.marca), cl.modelo.nombre,
                str(cl.ram) if cl.ram else "",
                str(cl.almacenamiento) if cl.almacenamiento else "",
                cl.get_estado_display(), _asignado_nombre(cl),
            ]
            for cl in qs
        ]
        return _pdf_response(
            "Reporte de Inventario — Celulares",
            headers, rows, "celulares_inventario.pdf",
        )


# ═══════════════════════════════════════════════
#  EQUIPOS — MONITOR
# ═══════════════════════════════════════════════

class MonitorViewSet(RolCCFilterMixin, viewsets.ModelViewSet):
    """Gestion de Monitores con exportacion."""
    cc_filter_field = "centro_costo_id"
    queryset = Monitor.objects.select_related(
        "marca", "modelo", "modelo__marca",
        "computador", "centro_costo",
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["estado", "marca", "modelo", "centro_costo", "computador"]
    search_fields = ["numero_inventario", "numero_serie", "marca__nombre", "modelo__nombre"]
    ordering_fields = ["fecha_ingreso", "estado", "marca__nombre"]

    def get_serializer_class(self):
        """Usa MonitorReadSerializer (anidado) para GET y MonitorWriteSerializer (IDs planos) para escritura."""
        if self.action in ("list", "retrieve"):
            return MonitorReadSerializer
        return MonitorWriteSerializer

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Genera y descarga un archivo Excel con todos los monitores del queryset filtrado."""
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monitores"
        ws.append([
            "N Inventario", "Marca", "Modelo", "N Serie",
            "Estado", "Computador", "Asignado a", "Area",
            "Fecha Ingreso", "Notas",
        ])
        for mn in qs:
            ws.append([
                mn.numero_inventario or "", str(mn.marca or ""),
                str(mn.modelo or ""), mn.numero_serie or "",
                mn.get_estado_display(),
                str(mn.computador) if mn.computador else "",
                _asignado_nombre(mn), _asignado_area(mn),
                str(mn.fecha_ingreso) if mn.fecha_ingreso else "",
                mn.notas,
            ])
        return _excel_response(ws, wb, "monitores_inventario.xlsx")


# ═══════════════════════════════════════════════
#  SOFTWARE INSTALADO
# ═══════════════════════════════════════════════

class ComputadorSoftwareViewSet(viewsets.ModelViewSet):
    """Software instalado en computadores."""
    queryset = ComputadorSoftware.objects.select_related(
        "computador", "software",
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["computador", "software"]
    search_fields = ["software__nombre", "version"]
    pagination_class = None

    def get_serializer_class(self):
        """Selecciona serializer de lectura (anidado) o escritura (IDs planos) según la acción."""
        if self.action in ("list", "retrieve"):
            return ComputadorSoftwareReadSerializer
        return ComputadorSoftwareWriteSerializer


# ═══════════════════════════════════════════════
#  VISTAS EXTRA (Dashboard + Historial Global)
# ═══════════════════════════════════════════════

@api_view(["GET"])
def dashboard_stats(request):
    """
    Endpoint que agrega las estadísticas del sistema para el Dashboard.
    Retorna en un solo request:
    - Totales globales por tipo de equipo y por estado
    - Distribución de computadores por marca, tipo y estado
    - Distribución de celulares por marca y estado
    - Top 10 áreas con más equipos asignados
    - Equipos sin asignar (en bodega sin empleado)
    - Conteos de empleados activos y áreas
    Utilizado por la página Dashboard.jsx con Recharts.
    """
    total_computadores = Computador.objects.count()
    total_celulares = Celular.objects.count()
    total_monitores = Monitor.objects.count()
    total_chips = Chip.objects.count()

    comp_estado = dict(
        Computador.objects.values_list("estado")
        .annotate(c=Count("id")).values_list("estado", "c")
    )
    cel_estado = dict(
        Celular.objects.values_list("estado")
        .annotate(c=Count("id")).values_list("estado", "c")
    )
    mon_estado = dict(
        Monitor.objects.values_list("estado")
        .annotate(c=Count("id")).values_list("estado", "c")
    )
    chip_estado = dict(
        Chip.objects.values_list("estado")
        .annotate(c=Count("id")).values_list("estado", "c")
    )

    comp_marca = list(
        Computador.objects.values("marca__nombre")
        .annotate(count=Count("id")).order_by("-count")[:8]
    )
    cel_marca = list(
        Celular.objects.values("marca__nombre")
        .annotate(count=Count("id")).order_by("-count")[:8]
    )

    # Computadores por tipo (laptop, desktop, etc.)
    comp_tipo = list(
        Computador.objects.values("tipo_equipo")
        .annotate(count=Count("id")).order_by("-count")
    )

    sin_asignar = (
        Computador.objects.filter(empleado_asignado__isnull=True, estado="EN_BODEGA").count()
        + Celular.objects.filter(empleado_asignado__isnull=True, estado="EN_BODEGA").count()
        + Monitor.objects.filter(centro_costo__isnull=True, estado="EN_BODEGA").count()
    )

    # Global counts cross-type
    total_en_uso = (
        comp_estado.get("EN_USO", 0)
        + cel_estado.get("EN_USO", 0)
        + mon_estado.get("EN_USO", 0)
    )
    total_en_reparacion = (
        comp_estado.get("EN_REPARACION", 0)
        + cel_estado.get("EN_REPARACION", 0)
        + mon_estado.get("EN_REPARACION", 0)
    )
    total_pendiente_devolucion = (
        comp_estado.get("PENDIENTE_DEVOLUCION", 0)
        + cel_estado.get("PENDIENTE_DEVOLUCION", 0)
        + mon_estado.get("PENDIENTE_DEVOLUCION", 0)
    )
    total_de_baja = (
        comp_estado.get("DE_BAJA", 0)
        + cel_estado.get("DE_BAJA", 0)
        + mon_estado.get("DE_BAJA", 0)
    )

    # Equipment by area (areas with most active equipment)
    equipos_por_area = list(
        Computador.objects.filter(empleado_asignado__isnull=False)
        .values("empleado_asignado__area__nombre")
        .annotate(computadores=Count("id"))
        .order_by("-computadores")[:10]
    )
    # Enrich with celulares per area
    cel_por_area = dict(
        Celular.objects.filter(empleado_asignado__isnull=False)
        .values_list("empleado_asignado__area__nombre")
        .annotate(c=Count("id"))
        .values_list("empleado_asignado__area__nombre", "c")
    )
    for item in equipos_por_area:
        area = item["empleado_asignado__area__nombre"]
        item["celulares"] = cel_por_area.get(area, 0)
        item["total"] = item["computadores"] + item["celulares"]
        item["area"] = area
    # Sort by total
    equipos_por_area.sort(key=lambda x: x["total"], reverse=True)

    return Response({
        "total_computadores": total_computadores,
        "total_celulares": total_celulares,
        "total_monitores": total_monitores,
        "total_chips": total_chips,
        "total_equipos": total_computadores + total_celulares + total_monitores,
        "equipos_sin_asignar": sin_asignar,
        "total_empleados": Empleado.objects.filter(activo=True).count(),
        "total_areas": Area.objects.count(),
        # Global health
        "total_en_uso": total_en_uso,
        "total_en_reparacion": total_en_reparacion,
        "total_pendiente_devolucion": total_pendiente_devolucion,
        "total_de_baja": total_de_baja,
        "total_en_bodega": (
            comp_estado.get("EN_BODEGA", 0)
            + cel_estado.get("EN_BODEGA", 0)
            + mon_estado.get("EN_BODEGA", 0)
        ),
        # Per type
        "computadores_por_estado": {
            "EN_USO": comp_estado.get("EN_USO", 0),
            "EN_BODEGA": comp_estado.get("EN_BODEGA", 0),
            "EN_REPARACION": comp_estado.get("EN_REPARACION", 0),
            "PENDIENTE_DEVOLUCION": comp_estado.get("PENDIENTE_DEVOLUCION", 0),
            "DE_BAJA": comp_estado.get("DE_BAJA", 0),
        },
        "celulares_por_estado": {
            "EN_USO": cel_estado.get("EN_USO", 0),
            "EN_BODEGA": cel_estado.get("EN_BODEGA", 0),
            "EN_REPARACION": cel_estado.get("EN_REPARACION", 0),
            "PENDIENTE_DEVOLUCION": cel_estado.get("PENDIENTE_DEVOLUCION", 0),
            "DE_BAJA": cel_estado.get("DE_BAJA", 0),
        },
        "monitores_por_estado": {
            "EN_USO": mon_estado.get("EN_USO", 0),
            "EN_BODEGA": mon_estado.get("EN_BODEGA", 0),
            "EN_REPARACION": mon_estado.get("EN_REPARACION", 0),
            "PENDIENTE_DEVOLUCION": mon_estado.get("PENDIENTE_DEVOLUCION", 0),
            "DE_BAJA": mon_estado.get("DE_BAJA", 0),
        },
        "chips_por_estado": {
            "EN_USO": chip_estado.get("EN_USO", 0),
            "EN_BODEGA": chip_estado.get("EN_BODEGA", 0),
            "SUSPENDIDA": chip_estado.get("SUSPENDIDA", 0),
            "DE_BAJA": chip_estado.get("DE_BAJA", 0),
        },
        "computadores_por_marca": comp_marca,
        "celulares_por_marca": cel_marca,
        "computadores_por_tipo": comp_tipo,
        "equipos_por_area": equipos_por_area,
    })


@api_view(["GET"])
def historial_global(request):
    """
    Devuelve el historial de cambios combinado de todos los modelos con auditoría.
    Acepta ?tipo=computador|celular|chip para filtrar por tipo de equipo.
    Sin parámetro muestra los últimos 150 registros de cada tipo, mezclados y ordenados por fecha.
    Utilizado por la página Historial.jsx.
    """
    tipo = request.query_params.get("tipo", None)
    results = []

    if tipo is None or tipo == "computador":
        results += _serialize_history(
            Computador.history.all().select_related("history_user"), "Computador", 150
        )
    if tipo is None or tipo == "celular":
        results += _serialize_history(
            Celular.history.all().select_related("history_user"), "Celular", 150
        )

    # Ordenar desde el cambio mas reciente al mas antiguo
    results.sort(key=lambda x: x["fecha"], reverse=True)
    return Response(results[:200])


# ═══════════════════════════════════════════════
#  SOLICITUDES DE CORREO ELECTRÓNICO
# ═══════════════════════════════════════════════

class SolicitudCorreoViewSet(viewsets.ModelViewSet):
    """
    CRUD completo para solicitudes de alta/baja de correo electrónico.
    - GET /solicitudes-correo/          → lista con filtros
    - POST /solicitudes-correo/         → crea solicitud (dispara email automático)
    - PATCH /solicitudes-correo/{id}/   → actualiza estado / ticket CAU
    - POST /solicitudes-correo/{id}/cambiar_estado/ → cambia estado con validación
    - GET /solicitudes-correo/resumen/  → contadores por estado para el dashboard
    Filtros disponibles: ?tipo=ALTA|BAJA  ?estado=PENDIENTE|ENVIADO_CAU|COMPLETADO|RECHAZADO
    """
    queryset = SolicitudCorreo.objects.select_related(
        "centro_costo", "centro_costo__area", "solicitante"
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["tipo", "estado", "centro_costo"]
    search_fields    = ["empleado_nombre", "empleado_rut", "numero_ticket_cau"]
    ordering_fields  = ["created_at", "fecha_requerida", "estado"]
    ordering         = ["-created_at"]

    def get_serializer_class(self):
        """Usa el serializer de lectura (nested) para GET y el de escritura (IDs) para POST/PUT/PATCH."""
        if self.action in ("list", "retrieve"):
            return SolicitudCorreoReadSerializer
        return SolicitudCorreoWriteSerializer

    @action(detail=True, methods=["post"], url_path="cambiar_estado")
    def cambiar_estado(self, request, pk=None):
        """
        Cambia el estado de una solicitud con validación de transiciones permitidas.
        Body: { "estado": "ENVIADO_CAU"|"COMPLETADO"|"RECHAZADO", "numero_ticket_cau": "..." }
        Transiciones válidas:
          PENDIENTE   → ENVIADO_CAU | RECHAZADO
          ENVIADO_CAU → COMPLETADO  | RECHAZADO
        """
        solicitud   = self.get_object()
        nuevo_estado = request.data.get("estado")
        ticket_cau   = request.data.get("numero_ticket_cau", "")

        TRANSICIONES = {
            EstadoSolicitudCorreo.PENDIENTE:   [EstadoSolicitudCorreo.ENVIADO_CAU, EstadoSolicitudCorreo.RECHAZADO],
            EstadoSolicitudCorreo.ENVIADO_CAU: [EstadoSolicitudCorreo.COMPLETADO,  EstadoSolicitudCorreo.RECHAZADO],
        }

        permitidos = TRANSICIONES.get(solicitud.estado, [])
        if nuevo_estado not in permitidos:
            return Response(
                {"error": f"No se puede pasar de '{solicitud.estado}' a '{nuevo_estado}'."},
                status=400,
            )

        solicitud.estado = nuevo_estado
        if ticket_cau:
            solicitud.numero_ticket_cau = ticket_cau
        solicitud.save()
        return Response(SolicitudCorreoReadSerializer(solicitud).data)

    @action(detail=False, methods=["get"], url_path="resumen")
    def resumen(self, request):
        """
        Retorna contadores agrupados por estado para el dashboard de solicitudes.
        Respuesta: { "PENDIENTE": N, "ENVIADO_CAU": N, "COMPLETADO": N, "RECHAZADO": N, "total": N }
        """
        qs = SolicitudCorreo.objects.all()
        conteo = {e: qs.filter(estado=e).count() for e in EstadoSolicitudCorreo.values}
        conteo["total"] = qs.count()
        return Response(conteo)

    @action(detail=False, methods=["get"], url_path="exportar_excel")
    def exportar_excel(self, request):
        """Exporta todas las solicitudes de correo a un archivo Excel."""
        qs = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solicitudes Correo"

        headers = [
            "ID", "Tipo", "Estado", "Empleado", "RUT", "Cargo",
            "Centro de Costo", "Fecha Requerida", "N° Ticket CAU",
            "Registrado por", "Fecha Registro", "Observaciones",
        ]
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, s in enumerate(qs, 2):
            ws.append([
                s.id,
                s.get_tipo_display(),
                s.get_estado_display(),
                s.empleado_nombre,
                s.empleado_rut,
                s.empleado_cargo,
                str(s.centro_costo),
                s.fecha_requerida.strftime("%d/%m/%Y") if s.fecha_requerida else "",
                s.numero_ticket_cau,
                s.solicitante.get_full_name() or s.solicitante.username,
                s.created_at.strftime("%d/%m/%Y %H:%M") if s.created_at else "",
                s.observaciones,
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(cell.value or "")) for cell in col
            ) + 4

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="solicitudes_correo.xlsx"'
        return resp
