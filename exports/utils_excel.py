from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ======================================================================
# Estilos reutilizables
# ======================================================================
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

STATUS_COLORS = {
    'EN_USO': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'EN_BODEGA': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'EN_REPARACION': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'PENDIENTE_DEVOLUCION': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'DE_BAJA': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

ACCION_COLORS = {
    'CREACION': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'ACTUALIZACION': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'CAMBIO_ESTADO': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'ELIMINACION': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}


def _apply_header_style(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_adjust(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _write_title(ws, title, merge_end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
    cell = ws.cell(row=1, column=1)
    cell.value = f'{title} — Generado: {datetime.now():%d/%m/%Y %H:%M}'
    cell.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    cell.alignment = Alignment(horizontal='center')


def _add_summary_sheet(wb, title, total, en_uso, en_reparacion, de_baja):
    ws = wb.create_sheet('Resumen')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    data = [
        ('Total de equipos:', total),
        ('En uso:', en_uso),
        ('En reparacion:', en_reparacion),
        ('De baja:', de_baja),
    ]
    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    _auto_adjust(ws)


# ======================================================================
# Computadores
# ======================================================================
def generar_excel_computadores(computadores):
    """Recibe un queryset de Computador y devuelve un BytesIO con el .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Computadores'

    headers = [
        '#', 'N Inventario', 'N Serie', 'Tipo', 'Marca', 'Modelo',
        'Procesador', 'RAM', 'Almacenamiento', 'S.O.', 'Estado', 'Fecha Ingreso',
    ]
    _write_title(ws, 'Inventario de Computadores', len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    for idx, eq in enumerate(computadores, 1):
        row = idx + 3
        values = [
            idx,
            eq.numero_inventario,
            eq.numero_serie,
            eq.get_tipo_equipo_display(),
            str(eq.marca),
            eq.modelo.nombre if eq.modelo else '',
            str(eq.procesador) if eq.procesador else '',
            str(eq.ram) if eq.ram else '',
            str(eq.almacenamiento) if eq.almacenamiento else '',
            str(eq.sistema_operativo) if eq.sistema_operativo else '',
            eq.get_estado_display(),
            eq.fecha_ingreso.strftime('%d/%m/%Y') if eq.fecha_ingreso else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN
            if col == 11:
                fill = STATUS_COLORS.get(eq.estado)
                if fill:
                    cell.fill = fill

    _auto_adjust(ws)

    eq_list = list(computadores)
    _add_summary_sheet(
        wb,
        'Resumen — Computadores',
        len(eq_list),
        sum(1 for e in eq_list if e.estado == 'EN_USO'),
        sum(1 for e in eq_list if e.estado == 'EN_REPARACION'),
        sum(1 for e in eq_list if e.estado == 'DE_BAJA'),
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ======================================================================
# Celulares
# ======================================================================
def generar_excel_celulares(celulares):
    """Recibe un queryset de Celular y devuelve un BytesIO con el .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Celulares'

    headers = [
        '#', 'N Linea', 'IMEI', 'N Serie', 'Tipo', 'Marca', 'Modelo',
        'RAM', 'Almacenamiento', 'Estado', 'Fecha Ingreso',
    ]
    _write_title(ws, 'Inventario de Celulares', len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    for idx, cel in enumerate(celulares, 1):
        row = idx + 3
        values = [
            idx,
            cel.numero_linea or '',
            cel.imei or '',
            cel.numero_serie or '',
            cel.get_tipo_equipo_display(),
            str(cel.marca),
            cel.modelo.nombre if cel.modelo else '',
            str(cel.ram) if cel.ram else '',
            str(cel.almacenamiento) if cel.almacenamiento else '',
            cel.get_estado_display(),
            cel.fecha_ingreso.strftime('%d/%m/%Y') if cel.fecha_ingreso else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN
            if col == 10:
                fill = STATUS_COLORS.get(cel.estado)
                if fill:
                    cell.fill = fill

    _auto_adjust(ws)

    cel_list = list(celulares)
    _add_summary_sheet(
        wb,
        'Resumen — Celulares',
        len(cel_list),
        sum(1 for c in cel_list if c.estado == 'EN_USO'),
        sum(1 for c in cel_list if c.estado == 'EN_REPARACION'),
        sum(1 for c in cel_list if c.estado == 'DE_BAJA'),
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ======================================================================
# Historial de cambios
# ======================================================================
def generar_excel_historial(registros):
    """Recibe un queryset de RegistroCambio y devuelve un BytesIO con el .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Historial'

    headers = [
        '#', 'Fecha', 'Tipo Equipo', 'ID Equipo', 'Accion',
        'Campo', 'Valor Anterior', 'Valor Nuevo', 'Usuario', 'IP',
    ]
    _write_title(ws, 'Historial de Cambios', len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    for idx, reg in enumerate(registros, 1):
        row = idx + 3
        values = [
            idx,
            reg.fecha.strftime('%d/%m/%Y %H:%M') if reg.fecha else '',
            reg.content_type.model.capitalize() if reg.content_type else '',
            reg.object_id,
            reg.get_accion_display(),
            reg.campo or '',
            reg.valor_anterior or '',
            reg.valor_nuevo or '',
            reg.usuario or '',
            reg.direccion_ip or '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN
            if col == 5:
                fill = ACCION_COLORS.get(reg.accion)
                if fill:
                    cell.fill = fill

    _auto_adjust(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
